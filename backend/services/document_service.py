import os
import shutil
import tempfile
from typing import Optional, List, Dict, Any
from sqlalchemy import select
from sqlalchemy.orm import Session
from backend.config import settings
from backend.utils.llm_provider import LLMProvider
from backend.workers.tasks import process_document_task

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")

class DocumentService:
    def __init__(self, db: Session, tenant_id: str):
        self.db = db
        self.tenant_id = tenant_id
        self.llm_provider = LLMProvider(settings.OPENAI_API_KEY)

    async def retrieve_relevant_chunks(self, query: str, document_ids: Optional[List[str]] = None) -> list:
        from backend.workers.tasks import generate_openai_embeddings
        from backend.models import DocumentChunk
        query_vector = generate_openai_embeddings(query)
        distance_expr = DocumentChunk.embedding.cosine_distance(query_vector).label("distance")
        stmt = select(DocumentChunk, distance_expr).where(DocumentChunk.tenant_id == self.tenant_id)
        if document_ids:
            stmt = stmt.where(DocumentChunk.document_name.in_(document_ids))
        stmt = stmt.order_by("distance").limit(5)
        results = self.db.execute(stmt).all()
        chunks = []
        for chunk_obj, distance in results:
            chunks.append({
                "content": chunk_obj.content,
                "document_name": chunk_obj.document_name,
                "chunk_index": chunk_obj.chunk_index,
                "distance": float(distance or 0.0)
            })
        return chunks

    def upload_document(self, filename: str, file_obj) -> Dict[str, Any]:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        file_path = os.path.join(UPLOAD_DIR, f"{self.tenant_id}_{filename}")
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file_obj, buffer)
        task = process_document_task.apply_async(
            args=[self.tenant_id, file_path, filename],
            queue="ai_task_queue"
        )
        return {
            "status": "queued",
            "task_id": task.id,
            "document_name": filename
        }

    def get_documents(self) -> List[Dict[str, Any]]:
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        files = []
        prefix = f"{self.tenant_id}_"
        for filename in os.listdir(UPLOAD_DIR):
            if filename.startswith(prefix):
                display_name = filename[len(prefix):]
                path = os.path.join(UPLOAD_DIR, filename)
                files.append({
                    "filename": display_name,
                    "size": os.path.getsize(path),
                    "created_at": os.path.getmtime(path)
                })
        return files

    async def execute_query(self, query: str, document_ids: Optional[List[str]]) -> Dict[str, Any]:
        chunks = await self.retrieve_relevant_chunks(query, document_ids)
        answer = await self.llm_provider.complete(query, chunks)
        return {
            "answer": answer,
            "chunks": chunks
        }

    def export_document(self, format: str, content: str) -> Dict[str, Any]:
        if format.lower() == "excel":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Vendor Comparison Matrix"
            lines = content.split("\n")
            table_rows = []
            for line in lines:
                if "|" in line:
                    parts = [p.strip() for p in line.split("|")[1:-1]]
                    if parts and not all(p == "" or "-" in p for p in parts):
                        cleaned_parts = [p.replace("**", "").replace("`", "") for p in parts]
                        table_rows.append(cleaned_parts)
            for row_idx, row in enumerate(table_rows, start=1):
                for col_idx, val in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if row_idx == 1:
                        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.font = Font(name="Arial", size=10)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    thin = Side(border_style="thin", color="E4E4E7")
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = col[0].column_letter
                ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(temp_file.name)
            return {
                "file_path": temp_file.name,
                "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "filename": "vendor_comparison_matrix.xlsx"
            }
        elif format.lower() == "pdf":
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
            doc = SimpleDocTemplate(temp_file.name, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
            story = []
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'TitleStyle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=16,
                textColor=colors.HexColor('#18181b'),
                spaceAfter=15
            )
            body_style = ParagraphStyle(
                'BodyStyle',
                parent=styles['BodyText'],
                fontName='Helvetica',
                fontSize=9,
                leading=12,
                textColor=colors.HexColor('#27272a')
            )
            story.append(Paragraph("synq.to AI Document Analysis Report", title_style))
            story.append(Spacer(1, 10))
            lines = content.split("\n")
            table_rows = []
            text_paragraphs = []
            for line in lines:
                if "|" in line:
                    parts = [p.strip() for p in line.split("|")[1:-1]]
                    if parts and not all(p == "" or "-" in p for p in parts):
                        cleaned_parts = [p.replace("**", "").replace("`", "") for p in parts]
                        table_rows.append(cleaned_parts)
                else:
                    if line.strip():
                        text_paragraphs.append(line.strip())
            for tp in text_paragraphs[:2]:
                story.append(Paragraph(tp, body_style))
                story.append(Spacer(1, 6))
            if table_rows:
                col_widths = [150, 180, 180]
                formatted_table_data = []
                for r_idx, row in enumerate(table_rows):
                    formatted_row = []
                    for cell_val in row:
                        cell_style = ParagraphStyle(
                            'Cell',
                            parent=body_style,
                            textColor=colors.white if r_idx == 0 else colors.HexColor('#27272a'),
                            fontName='Helvetica-Bold' if r_idx == 0 else 'Helvetica'
                        )
                        formatted_row.append(Paragraph(cell_val, cell_style))
                    formatted_table_data.append(formatted_row)
                rl_table = RLTable(formatted_table_data, colWidths=col_widths)
                rl_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#18181B')),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
                ]))
                story.append(Spacer(1, 10))
                story.append(rl_table)
                story.append(Spacer(1, 10))
            for tp in text_paragraphs[2:]:
                story.append(Paragraph(tp, body_style))
                story.append(Spacer(1, 6))
            doc.build(story)
            return {
                "file_path": temp_file.name,
                "media_type": "application/pdf",
                "filename": "vendor_comparison_report.pdf"
            }
        raise HTTPException(status_code=400, detail="Invalid export format specified. Choose 'excel' or 'pdf'.")
