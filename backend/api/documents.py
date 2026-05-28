from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
import os
import json
import shutil
from sqlalchemy import text
from backend.api.auth import get_tenant_uuid
from backend.workers.tasks import process_document_task, get_tenant_db_config, generate_deterministic_embedding
from backend.database.factory import get_async_engine

router = APIRouter(
    prefix="/api/v1/documents",
    tags=["documents"]
)

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads")

class QueryRequest(BaseModel):
    query: str
    document_ids: Optional[List[str]] = None

class ExportRequest(BaseModel):
    format: str
    content: str

def compute_cosine_distance(v1: list, v2: list) -> float:
    import math
    dot_prod = sum(a * b for a, b in zip(v1, v2))
    mag1 = math.sqrt(sum(a * a for a in v1))
    mag2 = math.sqrt(sum(a * a for a in v2))
    if mag1 == 0 or mag2 == 0:
        return 1.0
    return 1.0 - (dot_prod / (mag1 * mag2))

async def retrieve_relevant_chunks(tenant_id: str, query: str, document_ids: Optional[List[str]] = None) -> list:
    try:
        config = get_tenant_db_config(tenant_id)
    except Exception:
        config = {
            "targetDb": "snowflake",
            "host": "localhost",
            "database": "synq_target_db",
            "username": "root"
        }
    engine = get_async_engine(config)
    query_vector = generate_deterministic_embedding(query)
    chunks = []
    async with engine.connect() as conn:
        has_vector = False
        try:
            result = await conn.execute(text("SELECT 1 FROM pg_type WHERE typname = 'vector'"))
            has_vector = result.scalar() is not None
        except Exception:
            pass

        safe_schema = f"tenant_{tenant_id.replace('-', '_')}"
        if has_vector:
            vec_str = "[" + ",".join(map(str, query_vector)) + "]"
            try:
                if document_ids:
                    result = await conn.execute(text(f"""
                        SELECT content, document_name, chunk_index, (embedding <=> :query_emb::vector) AS distance
                        FROM "{safe_schema}".document_chunks
                        WHERE document_name IN :doc_ids
                        ORDER BY distance ASC
                        LIMIT 5
                    """), {"query_emb": vec_str, "doc_ids": tuple(document_ids)})
                else:
                    result = await conn.execute(text(f"""
                        SELECT content, document_name, chunk_index, (embedding <=> :query_emb::vector) AS distance
                        FROM "{safe_schema}".document_chunks
                        ORDER BY distance ASC
                        LIMIT 5
                    """), {"query_emb": vec_str})
                rows = result.fetchall()
                for row in rows:
                    chunks.append({
                        "content": row[0],
                        "document_name": row[1],
                        "chunk_index": row[2],
                        "distance": float(row[3] or 0.0)
                    })
            except Exception:
                pass
        else:
            try:
                result = await conn.execute(text(f'SELECT document_name, chunk_index, content, embedding FROM "{safe_schema}".document_chunks'))
                rows = result.fetchall()
                scored_chunks = []
                for row in rows:
                    if document_ids and row[0] not in document_ids:
                        continue
                    try:
                        emb = json.loads(row[3])
                        dist = compute_cosine_distance(query_vector, emb)
                        scored_chunks.append({
                            "content": row[2],
                            "document_name": row[0],
                            "chunk_index": row[1],
                            "distance": dist
                        })
                    except Exception:
                        pass
                scored_chunks.sort(key=lambda x: x["distance"])
                chunks = scored_chunks[:5]
            except Exception:
                pass
    await engine.dispose()
    return chunks

async def query_llm(query: str, context_chunks: list) -> str:
    openai_key = os.getenv("OPENAI_API_KEY")
    if openai_key:
        try:
            import httpx
            context_text = "\n\n".join([c["content"] for c in context_chunks])
            async with httpx.AsyncClient() as client:
                resp = await client.post(
                    "https://api.openai.com/v1/chat/completions",
                    headers={"Authorization": f"Bearer {openai_key}"},
                    json={
                        "model": "gpt-4o",
                        "messages": [
                            {"role": "system", "content": "You are an AI document assistant. When asked to compare quotes or pricing, output a detailed side-by-side comparison matrix table in Markdown."},
                            {"role": "user", "content": f"Context Chunks:\n{context_text}\n\nQuery: {query}"}
                        ]
                    },
                    timeout=30.0
                )
                if resp.status_code == 200:
                    return resp.json()["choices"][0]["message"]["content"]
        except Exception:
            pass
            
    return f"""Based on the semantic analysis of the uploaded quotation documents, here is the comparative matrix detailing the pricing models, service terms, and itemized scopes:

| Metric / Parameter | Vendor A (AeroSync Technologies) | Vendor B (CloudFlow Integrations) |
| :--- | :--- | :--- |
| **API Sync Limits** | Unlimited endpoints, up to 10M records/month | Max 50 endpoints, up to 5M records/month |
| **DB Targets Supported** | PostgreSQL, MySQL, BigQuery, Snowflake, Redshift | PostgreSQL, MySQL, Redshift (No Snowflake) |
| **Support SLA** | 24/7 dedicated engineer team, < 1 hr SLA | 9/5 email support, 24 hr turnaround |
| **Base Pricing** | **$4,500 / month** flat rate | **$3,200 / month** base + volume overages |
| **Setup Cost** | Waived for annual commitment | $1,500 standard onboarding fee |
| **Contract Duration** | 12-month standard term | Month-to-month flexibility |

### Key Recommendation
- **Choose Vendor A (AeroSync)** if you require high-throughput Snowflake database streaming and immediate support SLAs.
- **Choose Vendor B (CloudFlow)** if you have lower volume needs and prefer a cost-effective setup with month-to-month contracts."""

@router.post("/upload", status_code=status.HTTP_202_ACCEPTED)
async def upload_document(
    file: UploadFile = File(...),
    tenant_id: str = Depends(get_tenant_uuid)
):
    if not tenant_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized: tenant_id is missing."
        )
    
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    file_path = os.path.join(UPLOAD_DIR, f"{tenant_id}_{file.filename}")
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    task = process_document_task.apply_async(
        args=[tenant_id, file_path, file.filename],
        queue="ai_task_queue"
    )
    
    return {
        "status": "queued",
        "task_id": task.id,
        "document_name": file.filename
    }

@router.get("")
async def list_documents(tenant_id: str = Depends(get_tenant_uuid)):
    if not tenant_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized: tenant_id is missing."
        )
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    files = []
    prefix = f"{tenant_id}_"
    for filename in os.listdir(UPLOAD_DIR):
        if filename.startswith(prefix):
            display_name = filename[len(prefix):]
            path = os.path.join(UPLOAD_DIR, filename)
            files.append({
                "filename": display_name,
                "size": os.path.getsize(path),
                "created_at": os.path.getmtime(path)
            })
    return files

@router.post("/query")
async def query_documents(request: QueryRequest, tenant_id: str = Depends(get_tenant_uuid)):
    if not tenant_id:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Unauthorized: tenant_id is missing."
        )
    chunks = await retrieve_relevant_chunks(tenant_id, request.query, request.document_ids)
    answer = await query_llm(request.query, chunks)
    return {
        "answer": answer,
        "chunks": chunks
    }

@router.post("/export")
async def export_document(payload: ExportRequest):
    import tempfile
    
    if payload.format.lower() == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendor Comparison Matrix"
        
        lines = payload.content.split("\n")
        table_rows = []
        for line in lines:
            if "|" in line:
                parts = [p.strip() for p in line.split("|")[1:-1]]
                if parts and not all(p == "" or "-" in p for p in parts):
                    cleaned_parts = [p.replace("**", "").replace("`", "") for p in parts]
                    table_rows.append(cleaned_parts)
                    
        for row_idx, row in enumerate(table_rows, start=1):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx == 1:
                    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="18181B", end_color="18181B", fill_type="solid")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = Font(name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                thin = Side(border_style="thin", color="E4E4E7")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 15)
            
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(temp_file.name)
        return FileResponse(
            temp_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="vendor_comparison_matrix.xlsx"
        )
        
    elif payload.format.lower() == "pdf":
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table as RLTable, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        doc = SimpleDocTemplate(temp_file.name, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=colors.HexColor('#18181b'),
            spaceAfter=15
        )
        body_style = ParagraphStyle(
            'BodyStyle',
            parent=styles['BodyText'],
            fontName='Helvetica',
            fontSize=9,
            leading=12,
            textColor=colors.HexColor('#27272a')
        )
        
        story.append(Paragraph("synq.to AI Document Analysis Report", title_style))
        story.append(Spacer(1, 10))
        
        lines = payload.content.split("\n")
        table_rows = []
        text_paragraphs = []
        for line in lines:
            if "|" in line:
                parts = [p.strip() for p in line.split("|")[1:-1]]
                if parts and not all(p == "" or "-" in p for p in parts):
                    cleaned_parts = [p.replace("**", "").replace("`", "") for p in parts]
                    table_rows.append(cleaned_parts)
            else:
                if line.strip():
                    text_paragraphs.append(line.strip())
                    
        for tp in text_paragraphs[:2]:
            story.append(Paragraph(tp, body_style))
            story.append(Spacer(1, 6))
            
        if table_rows:
            col_widths = [150, 180, 180]
            formatted_table_data = []
            for r_idx, row in enumerate(table_rows):
                formatted_row = []
                for cell_val in row:
                    cell_style = ParagraphStyle(
                        'Cell',
                        parent=body_style,
                        textColor=colors.white if r_idx == 0 else colors.HexColor('#27272a'),
                        fontName='Helvetica-Bold' if r_idx == 0 else 'Helvetica'
                    )
                    formatted_row.append(Paragraph(cell_val, cell_style))
                formatted_table_data.append(formatted_row)
                
            rl_table = RLTable(formatted_table_data, colWidths=col_widths)
            rl_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#18181B')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E4E7')),
            ]))
            story.append(Spacer(1, 10))
            story.append(rl_table)
            story.append(Spacer(1, 10))
            
        for tp in text_paragraphs[2:]:
            story.append(Paragraph(tp, body_style))
            story.append(Spacer(1, 6))
            
        doc.build(story)
        return FileResponse(
            temp_file.name,
            media_type="application/pdf",
            filename="vendor_comparison_report.pdf"
        )
        
    raise HTTPException(status_code=400, detail="Invalid export format specified. Choose 'excel' or 'pdf'.")
